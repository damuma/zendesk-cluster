#!/usr/bin/env python3
"""Combina los CSV de extraer_socios_apoya.py en un único .xlsx formateado.

Lee data/socios_apoya/{socios,apoya}_{mantener,descartar}.csv y produce
data/socios_apoya/socios_apoya.xlsx con una pestaña por lista, encabezados
formateados, fila de cabecera fija, autofiltro y anchos de columna.

Las columnas contacto_1..N de los CSV se resumen a:
    primer_contacto, segundo_contacto, tercer_contacto, otras_interacciones
(el resto de interacciones, separadas por "; ").

Uso:
    python scripts/socios_apoya_a_excel.py
    python scripts/socios_apoya_a_excel.py --input-dir data/socios_apoya
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)


def load_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def contact_cols(row: dict) -> list[str]:
    """Devuelve las fechas contacto_N no vacías en orden."""
    out = []
    i = 1
    while f"contacto_{i}" in row:
        v = (row.get(f"contacto_{i}") or "").strip()
        if v:
            out.append(v)
        i += 1
    return out


def reshape(rows: list[dict], descartar: bool) -> tuple[list[str], list[list]]:
    n_col = "n_contactos_ventana" if descartar else "n_contactos"
    headers = ["email", "nº contactos en ventana", "primer contacto",
               "segundo contacto", "tercer contacto", "otras interacciones en ventana"]
    if descartar:
        headers.append("interacciones posteriores (motivo de descarte)")

    data = []
    for r in rows:
        fechas = contact_cols(r)
        primero = fechas[0] if len(fechas) > 0 else ""
        segundo = fechas[1] if len(fechas) > 1 else ""
        tercero = fechas[2] if len(fechas) > 2 else ""
        otras = "; ".join(fechas[3:]) if len(fechas) > 3 else ""
        row = [r.get("email", ""), int(r.get(n_col) or 0), primero, segundo, tercero, otras]
        if descartar:
            row.append(r.get("interacciones_posteriores", ""))
        data.append(row)
    return headers, data


def style_sheet(ws, headers: list[str], widths: list[int]) -> None:
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    # wrap en las columnas largas (otras interacciones / posteriores)
    for c in range(6, len(headers) + 1):
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")


# email, n, primero, segundo, tercero, otras, [posteriores]
WIDTHS = [38, 12, 18, 18, 18, 45, 70]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", default="data/socios_apoya")
    parser.add_argument("--output", default=None,
                        help="Ruta del .xlsx (default: <input-dir>/socios_apoya.xlsx)")
    args = parser.parse_args()

    indir = Path(args.input_dir)
    out = Path(args.output) if args.output else indir / "socios_apoya.xlsx"

    sheets = [
        ("socios — mantener", "socios_mantener.csv", False),
        ("socios — descartar", "socios_descartar.csv", True),
        ("apoya — mantener", "apoya_mantener.csv", False),
        ("apoya — descartar", "apoya_descartar.csv", True),
    ]

    wb = Workbook()
    resumen = wb.active
    resumen.title = "Resumen"

    counts = {}
    first = True
    for title, fname, descartar in sheets:
        rows = load_csv(indir / fname)
        counts[title] = len(rows)
        ws = wb.create_sheet(title=title)
        headers, data = reshape(rows, descartar)
        ws.append(headers)
        for d in data:
            ws.append(d)
        style_sheet(ws, headers, WIDTHS[:len(headers)])

    # Pestaña Resumen
    resumen["A1"] = "Remitentes a socios@ / apoya@"
    resumen["A1"].font = TITLE_FONT
    notas = [
        "",
        "Ventana: 4-mar-2026 a 8-abr-2026 (ambos incluidos). Fechas en horario de Madrid.",
        "«Mantener» = escribieron en la ventana y NO volvieron a contactar a socios ni apoya desde el 9-abr-2026.",
        "«Descartar» = volvieron a escribir a socios o apoya a partir del 9-abr (ver columna de interacciones posteriores).",
        "Excluidas direcciones internas @eldiario.es.",
        "",
        "Recuentos:",
    ]
    r = 2
    for line in notas:
        resumen.cell(row=r, column=1, value=line)
        r += 1
    for title, _f, _d in sheets:
        resumen.cell(row=r, column=1, value=title)
        resumen.cell(row=r, column=2, value=counts[title])
        r += 1
    resumen.column_dimensions["A"].width = 95
    resumen.column_dimensions["B"].width = 12

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"✅ Excel generado: {out}")
    for title, _f, _d in sheets:
        print(f"   {title}: {counts[title]} filas")


if __name__ == "__main__":
    main()
